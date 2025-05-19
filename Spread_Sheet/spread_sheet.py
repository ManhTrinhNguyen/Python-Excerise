
# Write a program that:

# - reads the provided spreadsheet file "employees.xlsx" with the following information/columns: "name", "years of experience", "job title", "date of birth"
  
# - creates a new spreadsheet file `"employees_sorted.xlsx"` with following info/columns: `"name"`, `"years of experience"`, where the years of experience is sorted in descending order: so the employee name with the most experience in years is on top.


import openpyxl
from openpyxl import Workbook

# Get the employees list from "./employees.xlsx"
employees_file = openpyxl.load_workbook("./employees.xlsx").active

# First task

## Loop through those spreadsheet role to get Name and others  start from second row . Bcs The First row is a Letter A 
## range() will create a list of number to iterate through 

## Logic of this is I want to loop through the Max Row of this Sheet and start from the Second row bcs  first row of the Sheet is Letters 
## Now I have iteration number from second to Max Row of the sheet I want to get a value of single colume in at specific row 
## Then I will set a variables for each of that colume 

## Get employee_file max row 
employees_file_row = employees_file.max_row

for employees_row in range(2, employees_file.max_row + 1):
  # I want to get name of all employees I will use cell(row, colume) to get specific colume I want and use .value to get value of  it 
  name_all_employees = employees_file.cell(employees_row, 1).value

  # Get Year of experience 
  years_of_experience = employees_file.cell(employees_row, 2).value

  # Get Job title 
  job_title = employees_file.cell(employees_row, 3).value

  # Get Date of birth 
  date_of_birth = employees_file.cell(employees_row, 4).value

# - creates a new spreadsheet file `"employees_sorted.xlsx"` with following info/columns: `"name"`, `"years of experience"`, where the years of experience is sorted in descending order: so the employee name with the most experience in years is on top.

## First I will create a new list of tuble that have name and year of experience 
new_employees_file = []

## Then Following the logic in the First task I will have name all employees and year of experience 

for employees_row in range(2, employees_file.max_row + 1):
  name_all_employees = employees_file.cell(employees_row, 1).value

  years_of_experience = employees_file.cell(employees_row, 2).value

  ## Now I will create a new tuple with name and experience
  my_tuple_of_name_and_experience = (name_all_employees, years_of_experience)

  ## I will append those tuples to my new list 
  new_employees_file.append(my_tuple_of_name_and_experience)

## Now I have my new list of employees file with year of experience and and 
## I will use bult-in sort function to sort in descending order of the year of experience 
##  lambda x: x[1] means: "For each item x in the list, sort it using the second element (x[1], which is years_of_experience)."
sorted_new_employees_file = sorted(new_employees_file, key=lambda x: x[1], reverse=True)

## Now I have a new sorted list of a employee I will creates a new spreadsheet file `"employees_sorted.xlsx"` with following info/columns: `"name"`, `"years of experience"`

## First I will create a new sheet . 
## To create a new sheet I need to import Workbook from openpxyl

### Create new work book 
wb = Workbook()

### Active work sheet
ws = wb.active

### Add headers 
ws.append(["Name", "Years of Experiences"])

## I have a new Workbook I will loop through the new sorted employee list

for employee in sorted_new_employees_file:
  ## Now I will add data rows
  ws.append([employee[0], employee[1]])

### Now a can save a new workbook 
wb.save("new_employees.xlsx")
